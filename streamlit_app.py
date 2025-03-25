import streamlit as st
import pandas as pd
import base64
from io import BytesIO
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tempfile
import os

def decode_and_insert_images(input_df, column_names):
    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_idx, col_name in enumerate(input_df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(input_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            col_name = input_df.columns[col_idx - 1]

            if col_name in column_names and isinstance(value, str) and value.strip():
                try:
                    image_data = base64.b64decode(value)
                    image = Image.open(BytesIO(image_data))
                    image.thumbnail((100, 100))
                    img_io = BytesIO()
                    image.save(img_io, format='PNG')
                    img_io.seek(0)
                    img = XLImage(img_io)
                    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                    ws.add_image(img, cell_ref)
                    ws.row_dimensions[row_idx].height = 80
                    ws.column_dimensions[get_column_letter(col_idx)].width = 18
                except Exception:
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    output_path = os.path.join(tempfile.gettempdir(), "output_with_images.xlsx")
    wb.save(output_path)
    return output_path

st.title("Base64 to Image Excel Converter")

uploaded_file = st.file_uploader("Upload an Excel or CSV file", type=["xlsx", "csv"])

if uploaded_file:
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    columns = df.columns.tolist()
    selected_columns = st.multiselect("Select the columns containing base64 images", columns)

    if st.button("Process File") and selected_columns:
        st.write("Processing file...")
        result_path = decode_and_insert_images(df, selected_columns)
        with open(result_path, "rb") as f:
            st.download_button(label="Download Excel with Images", data=f, file_name="Pricelist_with_images.xlsx")